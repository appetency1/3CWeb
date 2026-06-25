import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# 样式定义
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
subheader_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ========== Sheet 1: 轮播图 ==========
ws1 = wb.active
ws1.title = "轮播图"

headers1 = ["序号", "文件名", "主题"]
ws1.append(headers1)
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

banners = [
    ("banner-1.jpg", "新机首发 旗舰降临"),
    ("banner-2.jpg", "笔记本专场 效率升级"),
    ("banner-3.jpg", "影音盛宴 沉浸体验"),
    ("banner-4.jpg", "智能穿戴 健康随行"),
    ("banner-5.jpg", "摄影专区 记录精彩"),
    ("banner-6.jpg", "游戏装备 燃爆全场"),
    ("banner-7.jpg", "数码配件 一应俱全"),
    ("banner-8.jpg", "限时特惠 数码狂欢"),
]

for i, (fname, theme) in enumerate(banners, 1):
    ws1.append([i, fname, theme])
    for cell in ws1[i+1]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if cell.column == 1 else "left", vertical="center")

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 30

# ========== Sheet 2: 商品图 ==========
ws2 = wb.create_sheet(title="商品图")

headers2 = ["序号", "品类", "文件名", "对应商品"]
ws2.append(headers2)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

categories = [
    ("智能手机", [
        ("goods-apple-iphone15-pro-max.jpg", "iPhone 15 Pro Max 1TB"),
        ("goods-apple-iphone15-pro.jpg", "iPhone 15 Pro 256G"),
        ("goods-apple-iphone-se.jpg", "iPhone SE 第三代"),
        ("goods-huawei-mate60-pro.jpg", "HUAWEI Mate 60 Pro"),
        ("goods-xiaomi-14-ultra.jpg", "小米14 Ultra"),
        ("goods-samsung-s24-ultra.jpg", "Samsung Galaxy S24 Ultra"),
        ("goods-oppo-find-x7-ultra.jpg", "OPPO Find X7 Ultra"),
        ("goods-vivo-x100-pro.jpg", "vivo X100 Pro"),
        ("goods-oneplus-12.jpg", "一加 12"),
        ("goods-honor-magic6.jpg", "荣耀 Magic6 Pro"),
        ("goods-google-pixel9.jpg", "Google Pixel 9 Pro"),
        ("goods-nothing-phone3.jpg", "Nothing Phone 3"),
        ("goods-sony-xperia1.jpg", "索尼 Xperia 1 VI"),
        ("goods-zte-axon60.jpg", "中兴 Axon 60 Ultra"),
        ("goods-meizu-21pro.jpg", "魅族 21 Pro"),
        ("goods-asus-rog8.jpg", "ROG Phone 8 Pro"),
        ("goods-nokia-3210.jpg", "诺基亚 3210 复刻版"),
        ("goods-philips-e2620.jpg", "飞利浦 E2620 老人机"),
        ("goods-ktouch-g88.jpg", "天语 G88 4G 老人手机"),
        ("goods-agm-m7.jpg", "AGM M7 三防手机"),
    ]),
    ("笔记本电脑", [
        ("goods-apple-macbook-air-m3.jpg", "MacBook Air M3 13英寸"),
        ("goods-apple-macbook-pro-16.jpg", "MacBook Pro M3 Max 16英寸"),
        ("goods-lenovo-thinkpad-x1.jpg", "ThinkPad X1 Carbon Gen12"),
        ("goods-huawei-matebook-x-pro.jpg", "HUAWEI MateBook X Pro 2024"),
        ("goods-asus-rog-zephyrus-16.jpg", "ASUS ROG 幻16 2024"),
        ("goods-dell-xps-16.jpg", "Dell XPS 16"),
        ("goods-microsoft-surface-laptop-6.jpg", "Surface Laptop 6"),
        ("goods-xiaomi-redmibook-16.jpg", "RedmiBook Pro 16 2024"),
        ("goods-hp-spectre.jpg", "惠普 Spectre x360 14"),
        ("goods-honor-magicbook.jpg", "荣耀 MagicBook Pro 16"),
        ("goods-mechrevo-15pro.jpg", "机械革命 耀世15 Pro"),
        ("goods-razer-blade16.jpg", "雷蛇 灵刃16"),
        ("goods-lg-gram17.jpg", "LG Gram 17"),
        ("goods-huawei-matebook-d16.jpg", "华为 MateBook D16"),
        ("goods-hp-omni10.jpg", "惠普 暗影精灵10"),
        ("goods-lenovo-thinkbook16.jpg", "ThinkBook 16+ 2024"),
    ]),
    ("台式电脑", [
        ("goods-apple-imac-24.jpg", "iMac 24英寸 M3"),
        ("goods-apple-mac-studio.jpg", "Mac Studio M2 Ultra"),
        ("goods-lenovo-legion-7000k.jpg", "联想 拯救者 刃7000K"),
        ("goods-dell-optiplex-plus.jpg", "戴尔 OptiPlex Tower Plus"),
        ("goods-hp-desktop.jpg", "惠普 暗影精灵 台式机"),
        ("goods-asus-rog-desktop.jpg", "华硕 ROG 魔刃X"),
        ("goods-huawei-matests.jpg", "华为 MateStation S"),
        ("goods-lenovo-xiaoxin-mini.jpg", "联想 小新 Mini 2024"),
    ]),
    ("平板电脑", [
        ("goods-apple-ipad-pro-m4.jpg", "iPad Pro M4 13英寸 / 11英寸（共用）"),
        ("goods-apple-ipad-air-m2.jpg", "iPad Air M2"),
        ("goods-huawei-matepad-pro.jpg", "HUAWEI MatePad Pro 13.2"),
        ("goods-samsung-tab-s9-ultra.jpg", "Samsung Tab S9 Ultra"),
        ("goods-xiaomi-pad-6s-pro.jpg", "小米平板 6S Pro"),
        ("goods-lenovo-xiaoxin.jpg", "联想 小新 Pad Pro 12.7"),
        ("goods-honor-pad9.jpg", "荣耀 Pad 9"),
        ("goods-xiaomi-pad5.jpg", "小米 Pad 5 Pro 12.4"),
        ("goods-huawei-matepad-se.jpg", "华为 MatePad SE 11"),
    ]),
    ("耳机/音箱", [
        ("goods-sony-wh-1000xm5.jpg", "Sony WH-1000XM5"),
        ("goods-sony-wf1000xm5.jpg", "Sony WF-1000XM5"),
        ("goods-apple-airpods-pro-2.jpg", "AirPods Pro 2 USB-C"),
        ("goods-apple-airpods-max.jpg", "AirPods Max"),
        ("goods-bose-qc-ultra.jpg", "Bose QC Ultra 头戴式"),
        ("goods-samsung-buds2-pro.jpg", "Samsung Galaxy Buds2 Pro"),
        ("goods-huawei-freebuds-pro-3.jpg", "HUAWEI FreeBuds Pro 3"),
        ("goods-marshall-kilburn-ii.jpg", "Marshall Kilburn II"),
        ("goods-jbl-charge-5.jpg", "JBL Charge 5"),
        ("goods-jbl-flip7.jpg", "JBL Flip 7"),
        ("goods-sennheiser-momentum4.jpg", "森海塞尔 Momentum 4"),
        ("goods-beyerdynamic-freebyrd.jpg", "拜亚动力 Free BYRD"),
        ("goods-huawei-soundx.jpg", "华为 Sound X 2024"),
        ("goods-xiaomi-soundpro.jpg", "小米 Sound Pro"),
        ("goods-edifier-s2000.jpg", "漫步者 S2000MKIII"),
        ("goods-akg-n5005.jpg", "AKG N5005"),
    ]),
    ("相机/摄像机", [
        ("goods-sony-a7m4.jpg", "Sony A7M4 / A7C2（共用）"),
        ("goods-canon-eos-r6-ii.jpg", "佳能 EOS R6 Mark II"),
        ("goods-dji-osmo-pocket-3.jpg", "DJI Osmo Pocket 3"),
        ("goods-gopro-hero-12.jpg", "GoPro HERO 12 Black"),
        ("goods-fujifilm-instax-mini-12.jpg", "Fujifilm Instax Mini 12"),
        ("goods-nikon-z6iii.jpg", "尼康 Z6 III"),
        ("goods-panasonic-s5ii.jpg", "松下 Lumix S5 II"),
        ("goods-dji-mini4pro.jpg", "大疆 Mini 4 Pro"),
        ("goods-dji-air3.jpg", "大疆 Air 3"),
    ]),
    ("智能手表/手环", [
        ("goods-apple-watch-s9.jpg", "Apple Watch Series 9 45mm"),
        ("goods-apple-watch-ultra-2.jpg", "Apple Watch Ultra 2"),
        ("goods-samsung-watch6-classic.jpg", "Galaxy Watch6 Classic 47mm"),
        ("goods-huawei-watch-gt-4.jpg", "HUAWEI Watch GT 4"),
        ("goods-garmin-fenix-7x-pro.jpg", "Garmin Fenix 7X Pro"),
        ("goods-amazfit-trex-3.jpg", "Amazfit T-Rex 3"),
        ("goods-xiaomi-band-9-pro.jpg", "小米手环 9 Pro"),
        ("goods-xiaomi-band-9.jpg", "小米手环 9 / HUAWEI Band 9 / Fitbit Charge 6（共用）"),
        ("goods-xiaomi-band8.jpg", "小米手环 8 Pro"),
        ("goods-oppo-band2.jpg", "OPPO 手环 2"),
        ("goods-huawei-band8.jpg", "华为手环 8"),
        ("goods-honor-band9.jpg", "荣耀手环 9"),
    ]),
    ("手机配件", [
        ("goods-anker-100w-gan.jpg", "Anker 100W 氮化镓充电器"),
        ("goods-samsung-t7-2tb.jpg", "三星 T7 移动固态硬盘 2TB"),
        ("goods-logitech-mx-master-3s.jpg", "罗技 MX Master 3S 鼠标"),
        ("goods-belkin-7in1-hub.jpg", "Belkin 7合1 拓展坞"),
        ("goods-ugreen-100w-cable.jpg", "绿联 100W 快充数据线 2米"),
        ("goods-ugreen-stand.jpg", "绿联 MagSafe 手机支架"),
        ("goods-spigen-iphone-case.jpg", "Spigen iPhone 15 磁吸保护壳"),
        ("goods-pitaka-case.jpg", "PITAKA 凯夫拉手机壳"),
        ("goods-xiaomi-powerbank.jpg", "小米 20000mAh 移动电源 3"),
        ("goods-baseus-65wgan.jpg", "倍思 65W 氮化镓充电器"),
        ("goods-baseus-cable.jpg", "倍思 100W 编织数据线"),
        ("goods-baseus-stand.jpg", "倍思 磁吸车载支架"),
        ("goods-romoss-30000.jpg", "罗马仕 30000mAh 充电宝"),
        ("goods-xiaomi-wireless55w.jpg", "小米 55W 无线充电器"),
        ("goods-apple-magsafe.jpg", "苹果 MagSafe 充电器"),
        ("goods-sandisk-extremepro.jpg", "闪迪 Extreme Pro 256G U盘"),
        ("goods-kingston-sd128.jpg", "金士顿 Canvas Select 128G SD卡"),
        ("goods-sandisk-ssd1t.jpg", "闪迪 1TB 移动固态"),
    ]),
    ("对讲机", [
        ("goods-xiaomi-talkie-3.jpg", "小米对讲机 3"),
        ("goods-xiaomi-talkie2s.jpg", "小米对讲机 2S"),
        ("goods-baofeng-uv5r.jpg", "宝锋 UV-5R"),
        ("goods-beifeng-bf518.jpg", "北峰 BF-518"),
    ]),
    ("网络设备", [
        ("goods-tplink-ax6000.jpg", "TP-Link AX6000"),
        ("goods-tplink-be65.jpg", "TP-Link BE65 WiFi7"),
        ("goods-xiaomi-be7000.jpg", "小米 BE7000 Wi-Fi 7"),
        ("goods-xiaomi-ax9000.jpg", "小米 AX9000 三频WiFi6"),
        ("goods-asus-rt-ax86u.jpg", "华硕 RT-AX86U"),
        ("goods-h3c-mini-s8gu.jpg", "H3C 千兆交换机"),
        ("goods-huawei-ax6.jpg", "华为 AX6 WiFi6"),
        ("goods-mercury-sg108p.jpg", "水星 SG108P PoE交换机"),
    ]),
]

row_idx = 2
seq = 1
for cat_name, items in categories:
    for fname, product in items:
        ws2.append([seq, cat_name, fname, product])
        for cell in ws2[row_idx]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if cell.column in (1, 2) else "left", vertical="center")
        row_idx += 1
        seq += 1

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 32
ws2.column_dimensions['D'].width = 40

# ========== Sheet 3: 汇总 ==========
ws3 = wb.create_sheet(title="汇总")

headers3 = ["品类", "图片数"]
ws3.append(headers3)
for cell in ws3[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

summary = [
    ("智能手机", 20),
    ("笔记本电脑", 16),
    ("台式电脑", 8),
    ("平板电脑", 9),
    ("耳机/音箱", 16),
    ("相机/摄像机", 9),
    ("智能手表/手环", 12),
    ("手机配件", 18),
    ("对讲机", 4),
    ("网络设备", 8),
    ("轮播图", 8),
    ("合计", 128),
]

for i, (cat, count) in enumerate(summary, 2):
    ws3.append([cat, count])
    for cell in ws3[i]:
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if cell.column == 2 else "left", vertical="center")
    if cat == "合计":
        for cell in ws3[i]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 12

# 保存
output_path = r"D:\AI Code\ClothesDemo\素材清单.xlsx"
wb.save(output_path)
print(f"Excel已生成: {output_path}")
